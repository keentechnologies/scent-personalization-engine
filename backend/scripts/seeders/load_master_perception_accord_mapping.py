import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.constants.accords import ACCORD_DISPLAY_NAMES, AccordKey
from app.constants.perception import PERCEPTION_DISPLAY_NAMES, PerceptionKey
from app.models.master_perception_table import MasterPerceptionTable
from app.models.master_accord_table import MasterAccordTable
from app.models.master_perception_accord_mapping import MasterPerceptionAccordMapping

DATABASE_URL = settings.DATABASE_URL
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(
    autocommit=False,
    autoflush=False,
    bind=engine,
)

SCORE_LEVEL_VALUES = {
    "HIGH": 0.75,
    "MOD-HIGH": 0.5,
    "MOD-LOW": 0.25,
    "LOW": 0.0
}

def load_master_perception_accord_mapping():
    session = SessionLocal()
    try:
        # Load workbook
        excel_file = "/app/Recomendation engine mapping.xlsx"
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb['6-6 cluster']

        # Construct lookup maps (normalize: lowercase, remove spaces and slashes)
        def normalize(name):
            return name.lower().replace(" ", "").replace("/", "") if name else ""

        norm_accords = {normalize(v): k.value for k, v in ACCORD_DISPLAY_NAMES.items()}
        norm_perceptions = {normalize(v): k.value for k, v in PERCEPTION_DISPLAY_NAMES.items()}

        # 1. Map columns (perception keys)
        col_to_perception = {}
        for col in range(2, 26):
            h = ws.cell(row=1, column=col).value
            if h is not None:
                norm_h = normalize(h)
                p_key = norm_perceptions.get(norm_h)
                if p_key:
                    col_to_perception[col] = p_key

        # Delete existing mappings to avoid duplicates
        session.query(MasterPerceptionAccordMapping).delete()
        session.commit()

        records = []
        # 2. Iterate rows (accords)
        for row in range(2, 26):
            accord_name = ws.cell(row=row, column=1).value
            if not accord_name:
                continue
            norm_a = normalize(accord_name)
            a_key = norm_accords.get(norm_a)
            if not a_key:
                print(f"Skipping unknown accord: {accord_name}")
                continue

            for col, p_key in col_to_perception.items():
                level = ws.cell(row=row, column=col).value
                level_str = str(level).strip().upper() if level is not None else ""
                val = SCORE_LEVEL_VALUES.get(level_str, 0.0)

                records.append(
                    MasterPerceptionAccordMapping(
                        accord_key=a_key,
                        perception_key=p_key,
                        score_level=level_str if level_str in SCORE_LEVEL_VALUES else "LOW",
                        score_value=val
                    )
                )

        session.bulk_save_objects(records)
        session.commit()
        print(f"Successfully loaded {len(records)} perception-accord mappings into master_perception_accord_mapping")

    except Exception as error:
        session.rollback()
        print("FAILED TO LOAD MASTER PERCEPTION ACCORD MAPPING TABLE", str(error))
    finally:
        session.close()

if __name__ == "__main__":
    load_master_perception_accord_mapping()
