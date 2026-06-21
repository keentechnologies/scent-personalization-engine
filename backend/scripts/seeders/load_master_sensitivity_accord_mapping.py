import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.constants.accords import ACCORD_DISPLAY_NAMES, AccordKey
from app.constants.sensitivities import SENSITIVITY_DISPLAY_NAMES, SensitivityKey
from app.models.master_sensitivity_table import MasterSensitivityTable
from app.models.master_accord_table import MasterAccordTable
from app.models.master_sensitivity_accord_mapping import MasterSensitivityAccordMapping

DATABASE_URL = settings.DATABASE_URL
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(
    autocommit=False,
    autoflush=False,
    bind=engine,
)

def load_master_sensitivity_accord_mapping():
    session = SessionLocal()
    try:
        # Load Excel
        excel_file = "/app/Recomendation engine mapping.xlsx"
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb['senstivity issue accord mapping']

        # Normalization helper
        def normalize(name):
            return name.lower().replace(" ", "").replace("/", "").strip() if name else ""

        norm_accords = {normalize(v): k.value for k, v in ACCORD_DISPLAY_NAMES.items()}
        norm_sensitivities = {normalize(v): k.value for k, v in SENSITIVITY_DISPLAY_NAMES.items()}

        # 1. Parse Excel data into a map of sensitivity_key -> set of accord_keys
        flagged_map = {}
        for row in range(2, 10):
            sens_name = ws.cell(row=row, column=1).value
            if not sens_name:
                continue
            norm_s = normalize(sens_name)
            s_key = norm_sensitivities.get(norm_s)
            if not s_key:
                print(f"Skipping unknown sensitivity in Excel: {sens_name}")
                continue

            accords_text = ws.cell(row=row, column=2).value or ""
            flagged_accords = set()
            for part in accords_text.split(","):
                norm_part = normalize(part)
                if norm_part:
                    a_key = norm_accords.get(norm_part)
                    if a_key:
                        flagged_accords.add(a_key)
                    else:
                        print(f"Skipping unknown accord: '{part.strip()}' under sensitivity '{sens_name}'")
            flagged_map[s_key] = flagged_accords

        # Delete existing entries to prevent duplication
        session.query(MasterSensitivityAccordMapping).delete()
        session.commit()

        records = []
        # 2. Iterate all 5 sensitivities and 24 accords to generate all 120 rows
        for s_key in [k.value for k in SensitivityKey]:
            flagged_set = flagged_map.get(s_key, set())
            for a_key in [k.value for k in AccordKey]:
                is_flagged = 1 if a_key in flagged_set else 0
                records.append(
                    MasterSensitivityAccordMapping(
                        sensitivity_key=s_key,
                        accord_key=a_key,
                        is_flagged=is_flagged
                    )
                )

        session.bulk_save_objects(records)
        session.commit()
        print(f"Successfully loaded {len(records)} mappings into master_sensitivity_accord_mapping")

    except Exception as error:
        session.rollback()
        print("FAILED TO LOAD MASTER SENSITIVITY ACCORD MAPPING TABLE", str(error))
    finally:
        session.close()

if __name__ == "__main__":
    load_master_sensitivity_accord_mapping()
