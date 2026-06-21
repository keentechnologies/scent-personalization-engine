import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.constants.accords import ACCORD_DISPLAY_NAMES, AccordKey
from app.constants.climate import ClimateKey, CLIMATE_KEY_TO_EXCEL_COL

# Explicitly import referenced models for SQLAlchemy metadata compile
from app.models.master_accord_table import MasterAccordTable
from app.models.master_climate_table import MasterClimateTable
from app.models.master_climate_accord_mapping import MasterClimateAccordMapping

DATABASE_URL = settings.DATABASE_URL
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(
    autocommit=False,
    autoflush=False,
    bind=engine,
)

def load_master_climate_accord_mapping():
    session = SessionLocal()
    try:
        # Load Excel
        excel_file = "/app/Recomendation engine mapping.xlsx"
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb['climate score count of breaches']

        # Normalization helper
        def normalize(name):
            return name.lower().replace(" ", "").replace("/", "").strip() if name else ""

        norm_accords = {normalize(v): k.value for k, v in ACCORD_DISPLAY_NAMES.items()}

        # Read headers
        headers = []
        for c in range(1, 15):
            val = ws.cell(row=1, column=c).value
            if val:
                headers.append(val.strip())
            else:
                headers.append("")

        # 1. Parse Excel data
        # Mapping: accord_key -> { excel_column_name: breach_count }
        accord_data = {}
        for row in range(2, 40):
            accord_name = ws.cell(row=row, column=1).value
            if not accord_name:
                continue
            
            norm_a = normalize(accord_name)
            a_key = norm_accords.get(norm_a)
            if not a_key:
                print(f"Skipping unknown accord in Excel: {accord_name}")
                continue

            column_values = {}
            for col_idx in range(2, 10): # Columns 2 to 9
                header = headers[col_idx - 1]
                if not header:
                    continue
                cell_val = ws.cell(row=row, column=col_idx).value
                # Default to 0 if cell is empty or None
                breach_count = int(float(cell_val)) if cell_val is not None else 0
                column_values[header] = breach_count

            accord_data[a_key] = column_values

        # Delete existing entries to prevent duplication
        session.query(MasterClimateAccordMapping).delete()
        session.commit()

        records = []
        # 2. Iterate all 24 accords and 18 climate combinations to generate 432 rows
        for a_key in [k.value for k in AccordKey]:
            for c_key in [k.value for k in ClimateKey]:
                excel_col = CLIMATE_KEY_TO_EXCEL_COL.get(ClimateKey(c_key))
                
                breach_count = 0
                if a_key in accord_data and excel_col:
                    breach_count = accord_data[a_key].get(excel_col, 0)
                
                records.append(
                    MasterClimateAccordMapping(
                        accord_key=a_key,
                        climate_key=c_key,
                        breach_count=breach_count
                    )
                )

        session.bulk_save_objects(records)
        session.commit()
        print(f"Successfully loaded {len(records)} mappings into master_climate_accord_mapping")

    except Exception as error:
        session.rollback()
        print("FAILED TO LOAD MASTER CLIMATE ACCORD MAPPING TABLE", str(error))
    finally:
        session.close()

if __name__ == "__main__":
    load_master_climate_accord_mapping()
